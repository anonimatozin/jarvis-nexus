# -*- coding: utf-8 -*-
"""
JARVIS TURBO v2.0 - Central de Comando Unificada
Modo especial que transforma o Jarvis em um centro de inteligencia,
criacao e automacao completo.
"""

import os
import csv
import json
import psutil
from pathlib import Path
from datetime import datetime


class JarvisTurbo:
    """Modo Turbo - Central de Comando do Jarvis.
    
    Capacidades:
      - Analise de arquivos e pastas
      - Visao computacional (screenshots, imagens)
      - Criacao e edicao de codigo
      - Automacao de tarefas
      - Relatorios de sistema
      - Integracao com todas as ferramentas
    """

    def __init__(self, engine=None):
        self.engine = engine
        self._ativo = False
        self._historico = []
        self._stats = {
            "tarefas_executadas": 0,
            "arquivos_analisados": 0,
            "codigos_gerados": 0,
            "pesquisas_realizadas": 0,
        }
        print("[TURBO] Modo Turbo inicializado")

    @property
    def ativo(self):
        return self._ativo

    def ativar(self):
        """Ativa o modo Turbo."""
        self._ativo = True
        self._registrar("Modo Turbo ativado")
        print("[TURBO] Modo Turbo ATIVADO")
        return "Modo Turbo ativado. Central de comando pronta, Sir."

    def desativar(self):
        """Desativa o modo Turbo."""
        self._ativo = False
        self._registrar("Modo Turbo desativado")
        print("[TURBO] Modo Turbo DESATIVADO")
        return "Modo Turbo desativado."

    def _registrar(self, acao):
        """Registra acao no historico."""
        ts = datetime.now().strftime("%H:%M:%S")
        self._historico.append({"hora": ts, "acao": acao})
        if len(self._historico) > 50:
            self._historico = self._historico[-50:]

    # ═══ ANALISE DE SISTEMA ═══

    def status_completo(self):
        """Retorna status completo do sistema."""
        cpu = psutil.cpu_percent(interval=0.5)
        ram = psutil.virtual_memory()
        disco = psutil.disk_usage("C:\\")
        rede = psutil.net_io_counters()

        status = {
            "cpu_percent": cpu,
            "cpu_nucleos": psutil.cpu_count(),
            "ram_total_gb": round(ram.total / (1024**3), 1),
            "ram_usada_gb": round(ram.used / (1024**3), 1),
            "ram_percent": ram.percent,
            "disco_total_gb": round(disco.total / (1024**3), 1),
            "disco_livre_gb": round(disco.free / (1024**3), 1),
            "disco_percent": round(disco.percent, 1),
            "rede_enviado_mb": round(rede.bytes_sent / (1024**2), 1),
            "rede_recebido_mb": round(rede.bytes_recv / (1024**2), 1),
            "uptime": self._uptime(),
            "turbo_stats": self._stats.copy(),
        }
        self._registrar("Status completo consultado")
        return status

    def _uptime(self):
        """Tempo ligado."""
        try:
            boot = psutil.boot_time()
            delta = datetime.now().timestamp() - boot
            horas = int(delta // 3600)
            mins = int((delta % 3600) // 60)
            return f"{horas}h {mins}m"
        except Exception:
            return "desconhecido"

    def formatar_status(self):
        """Status formatado para fala/texto."""
        s = self.status_completo()
        linhas = [
            f"CPU: {s['cpu_percent']}% ({s['cpu_nucleos']} nucleos)",
            f"RAM: {s['ram_usada_gb']}GB de {s['ram_total_gb']}GB ({s['ram_percent']}%)",
            f"Disco: {s['disco_livre_gb']}GB livres de {s['disco_total_gb']}GB",
            f"Rede: +{s['rede_recebido_mb']}MB / -{s['rede_enviado_mb']}MB",
            f"Uptime: {s['uptime']}",
            f"Tarefas: {s['turbo_stats']['tarefas_executadas']} executadas",
        ]
        return " | ".join(linhas)

    # ═══ ANALISE DE ARQUIVOS ═══

    def analisar_pasta(self, caminho, profundidade=2):
        """Analisa estrutura de uma pasta."""
        pasta = Path(caminho)
        if not pasta.exists():
            return {"erro": f"Pasta nao existe: {caminho}"}

        resultado = {
            "pasta": str(pasta),
            "total_arquivos": 0,
            "total_pastas": 0,
            "por_tipo": {},
            "maiores": [],
            "recentes": [],
            "problemas": [],
        }

        try:
            itens = list(pasta.iterdir())
            resultado["total_pastas"] = len([i for i in itens if i.is_dir()])
            resultado["total_arquivos"] = len([i for i in itens if i.is_file()])

            # Contagem por tipo
            for item in itens:
                if item.is_file():
                    ext = item.suffix.lower() or "(sem extensao)"
                    resultado["por_tipo"][ext] = resultado["por_tipo"].get(ext, 0) + 1

            # Maiores arquivos
            arquivos = [i for i in itens if i.is_file()]
            arquivos.sort(key=lambda x: x.stat().st_size, reverse=True)
            resultado["maiores"] = [
                {"nome": f.name, "tamanho_mb": round(f.stat().st_size / (1024**2), 2)}
                for f in arquivos[:5]
            ]

            # Mais recentes
            arquivos_recentes = sorted(arquivos, key=lambda x: x.stat().st_mtime, reverse=True)
            resultado["recentes"] = [
                {"nome": f.name, "modificado": datetime.fromtimestamp(f.stat().st_mtime).strftime("%d/%m %H:%M")}
                for f in arquivos_recentes[:5]
            ]

            # Verificar problemas
            for f in arquivos:
                if f.stat().st_size == 0:
                    resultado["problemas"].append(f"{f.name}: arquivo vazio")
                if f.suffix.lower() in (".tmp", ".temp", ".bak"):
                    resultado["problemas"].append(f"{f.name}: arquivo temporario")

            self._stats["arquivos_analisados"] += 1
            self._registrar(f"Analisou pasta: {pasta.name}")

        except PermissionError:
            resultado["erro"] = "Sem permissao de acesso"
        except Exception as ex:
            resultado["erro"] = str(ex)

        return resultado

    def formatar_analise(self, dados):
        """Analise formatada para texto."""
        if dados.get("erro"):
            return f"Erro: {dados['erro']}"

        linhas = [
            f"Analise de: {Path(dados['pasta']).name}",
            f"Arquivos: {dados['total_arquivos']} | Pastas: {dados['total_pastas']}",
        ]

        if dados["por_tipo"]:
            tipos = sorted(dados["por_tipo"].items(), key=lambda x: -x[1])[:5]
            tipos_str = ", ".join(f"{ext}: {qtd}" for ext, qtd in tipos)
            linhas.append(f"Tipos: {tipos_str}")

        if dados["maiores"]:
            maiores = dados["maiores"][:3]
            maiores_str = ", ".join(f"{m['nome']} ({m['tamanho_mb']}MB)" for m in maiores)
            linhas.append(f"Maiores: {maiores_str}")

        if dados["problemas"]:
            linhas.append(f"Problemas encontrados: {len(dados['problemas'])}")
            for p in dados["problemas"][:3]:
                linhas.append(f"  - {p}")

        return " | ".join(linhas)

    # ═══ AUTOMACAO ═══

    def executar_rotina(self, nome):
        """Executa uma rotina pre-definida."""
        rotinas = {
            "trabalho": self._rotina_trabalho,
            "gamer": self._rotina_gamer,
            "noturno": self._rotina_noturno,
            "manha": self._rotina_manha,
            "organizar": self._rotina_organizar,
        }

        rotina = rotinas.get(nome.lower())
        if rotina:
            resultado = rotina()
            self._registrar(f"Rotina executada: {nome}")
            self._stats["tarefas_executadas"] += 1
            return resultado
        return f"Rotina '{nome}' nao encontrada. Disponiveis: {', '.join(rotinas.keys())}"

    def _rotina_trabalho(self):
        """Modo trabalho: abre apps, ajusta volume."""
        if self.engine and self.engine.app_launcher:
            self.engine.app_launcher.abrir("chrome")
            if self.engine.system_control:
                self.engine.system_control.set_volume(40)
        return "Modo trabalho ativado: Chrome aberto, volume em 40%."

    def _rotina_gamer(self):
        """Modo gamer: ajusta para jogos."""
        if self.engine and self.engine.system_control:
            self.engine.system_control.set_volume(80)
        return "Modo gamer: volume em 80%, pronto para jogar."

    def _rotina_noturno(self):
        """Modo noturno: reduz brilho e notificacoes."""
        if self.engine and self.engine.system_control:
            self.engine.system_control.set_brightness(30)
            self.engine.system_control.set_volume(20)
        return "Modo noturno: brilho reduzido, volume baixo."

    def _rotina_manha(self):
        """Modo manha: briefing completo."""
        partes = []
        agora = datetime.now()
        partes.append(f"Bom dia, Sir. Sao {agora.strftime('%H:%M')}.")

        if self.engine and hasattr(self.engine, 'luzes') and self.engine.luzes:
            self.engine.luzes.cena_acordar()
            partes.append("Luzes acesas.")

        return " ".join(partes)

    def _rotina_organizar(self):
        """Organiza downloads."""
        if self.engine and hasattr(self.engine, 'dev_agent') and self.engine.dev_agent:
            return self.engine.dev_agent.organizar_downloads()
        return "Dev Agent indisponivel."

    # ═══ HISTORICO ═══

    def obter_historico(self, limite=10):
        """Retorna historico de acoes."""
        return self._historico[-limite:]

    def formatar_historico(self):
        """Historico formatado."""
        if not self._historico:
            return "Nenhuma acao registrada."
        linhas = []
        for item in self._historico[-10:]:
            linhas.append(f"[{item['hora']}] {item['acao']}")
        return " | ".join(linhas)

    # ═══ LEITURA E ANALISE DE ARQUIVOS ═══

    def ler_arquivo(self, caminho, max_linhas=500):
        """Le o conteudo de um arquivo. Retorna dict com info e conteudo."""
        arq = Path(caminho)
        if not arq.exists():
            return {"erro": f"Arquivo nao encontrado: {caminho}"}

        ext = arq.suffix.lower()
        tamanho = arq.stat().st_size
        resultado = {
            "nome": arq.name,
            "extensao": ext,
            "tamanho_kb": round(tamanho / 1024, 1),
            "caminho": str(arq),
        }

        # Excel
        if ext in (".xlsx", ".xls"):
            return self._ler_excel(arq, resultado)

        # CSV
        if ext == ".csv":
            return self._ler_csv(arq, resultado, max_linhas)

        # JSON
        if ext == ".json":
            return self._ler_json(arq, resultado)

        # Imagem - retorna info basica
        if ext in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".svg"):
            resultado["tipo"] = "imagem"
            resultado["conteudo"] = f"[Imagem: {arq.name} ({tamanho // 1024}KB)]"
            return resultado

        # PDF
        if ext == ".pdf":
            resultado["tipo"] = "pdf"
            resultado["conteudo"] = f"[PDF: {arq.name} ({tamanho // 1024}KB)]"
            return resultado

        # Arquivos de texto e codigo
        try:
            with open(arq, "r", encoding="utf-8", errors="replace") as f:
                linhas = f.readlines()
            total = len(linhas)
            conteudo = "".join(linhas[:max_linhas])
            resultado["tipo"] = "texto"
            resultado["total_linhas"] = total
            resultado["linhas_mostradas"] = min(total, max_linhas)
            resultado["conteudo"] = conteudo
            if total > max_linhas:
                resultado["conteudo"] += f"\n\n... [{total - max_linhas} linhas restantes]"
            return resultado
        except Exception as e:
            return {"erro": f"Erro ao ler: {e}", **resultado}

    def _ler_excel(self, arq, resultado):
        """Le planilha Excel."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
            resultado["tipo"] = "excel"
            resultado["abas"] = wb.sheetnames
            dados = {}
            for nome_aba in wb.sheetnames:
                ws = wb[nome_aba]
                linhas = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 100:  # limita a 100 linhas por aba
                        break
                    linhas.append([str(c) if c is not None else "" for c in row])
                dados[nome_aba] = linhas
            resultado["conteudo"] = dados
            resultado["resumo"] = self._resumir_excel(dados)
            wb.close()
            self._stats["arquivos_analisados"] += 1
            self._registrar(f"Leu Excel: {arq.name}")
            return resultado
        except ImportError:
            resultado["erro"] = "openpyxl nao instalado. Instale: pip install openpyxl"
            resultado["tipo"] = "excel"
            return resultado
        except Exception as e:
            resultado["erro"] = f"Erro ao ler Excel: {e}"
            return resultado

    def _resumir_excel(self, dados):
        """Gera resumo de planilha Excel."""
        partes = []
        for aba, linhas in dados.items():
            if linhas:
                header = linhas[0]
                n_linhas = len(linhas) - 1  # desconta header
                partes.append(f"Aba '{aba}': {len(header)} colunas, {n_linhas} linhas")
                if len(linhas) > 1:
                    partes.append(f"  Colunas: {', '.join(str(h) for h in header[:8])}")
                    partes.append(f"  Primeira linha: {', '.join(str(c) for c in linhas[1][:8])}")
        return " | ".join(partes) if partes else "Planilha vazia"

    def _ler_csv(self, arq, resultado, max_linhas):
        """Le arquivo CSV."""
        try:
            with open(arq, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                linhas = []
                for i, row in enumerate(reader):
                    if i >= max_linhas:
                        break
                    linhas.append(row)
            resultado["tipo"] = "csv"
            resultado["total_linhas"] = len(linhas)
            resultado["linhas_mostradas"] = len(linhas)
            if linhas:
                resultado["colunas"] = linhas[0]
                resultado["conteudo"] = linhas
                # Resumo
                n_cols = len(linhas[0]) if linhas else 0
                n_rows = len(linhas) - 1 if len(linhas) > 1 else 0
                resultado["resumo"] = f"CSV: {n_cols} colunas, {n_rows} linhas de dados"
            self._stats["arquivos_analisados"] += 1
            self._registrar(f"Leu CSV: {arq.name}")
            return resultado
        except Exception as e:
            resultado["erro"] = f"Erro ao ler CSV: {e}"
            return resultado

    def _ler_json(self, arq, resultado):
        """Le arquivo JSON."""
        try:
            with open(arq, "r", encoding="utf-8") as f:
                dados = json.load(f)
            resultado["tipo"] = "json"
            if isinstance(dados, dict):
                resultado["chaves"] = list(dados.keys())[:20]
                resultado["total_chaves"] = len(dados)
            elif isinstance(dados, list):
                resultado["total_itens"] = len(dados)
            resultado["conteudo"] = json.dumps(dados, indent=2, ensure_ascii=False)[:5000]
            self._stats["arquivos_analisados"] += 1
            self._registrar(f"Leu JSON: {arq.name}")
            return resultado
        except Exception as e:
            resultado["erro"] = f"Erro ao ler JSON: {e}"
            return resultado

    def formatar_leitura(self, dados):
        """Formata resultado de leitura para texto."""
        if dados.get("erro"):
            return f"Erro: {dados['erro']}"

        partes = [f"Arquivo: {dados.get('nome', '?')} ({dados.get('tamanho_kb', 0)}KB)"]

        if dados.get("tipo") == "excel" and dados.get("resumo"):
            partes.append(dados["resumo"])
        elif dados.get("tipo") == "csv" and dados.get("resumo"):
            partes.append(dados["resumo"])
        elif dados.get("tipo") == "json":
            if dados.get("chaves"):
                partes.append(f"Chaves: {', '.join(dados['chaves'][:10])}")
            elif dados.get("total_itens") is not None:
                partes.append(f"Itens: {dados['total_itens']}")
        elif dados.get("tipo") == "texto":
            partes.append(f"Linhas: {dados.get('total_linhas', '?')}")
            if dados.get("conteudo"):
                # Mostra primeiras linhas
                linhas = dados["conteudo"].split("\n")[:10]
                partes.append("Inicio:\n" + "\n".join(linhas))

        return "\n".join(partes)


# Singleton
_instance = None

def get_turbo(engine=None):
    global _instance
    if _instance is None:
        _instance = JarvisTurbo(engine=engine)
    return _instance
