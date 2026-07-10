# modules/planilhas.py
"""
J.A.R.V.I.S. - Gerador de Planilhas v1.0
Gera Excel (.xlsx) e CSV a partir de dados do Jarvis.
"""

import csv
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False

import sys
sys.path.append(str(Path(__file__).resolve().parent.parent))
from utils.logger import setup_logger

logger = setup_logger("planilhas")

REPORTS_DIR = Path(__file__).resolve().parent.parent / "data" / "relatorios"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


class PlanilhasGenerator:
    """Gera planilhas Excel e CSV."""

    def __init__(self):
        self.reports_dir = REPORTS_DIR

    def _estilizar_excel(self, ws, titulo: str):
        """Aplica estilo basico ao worksheet."""
        # Titulo
        ws.merge_cells("A1:F1")
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Header row
        for cell in ws[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    def gerar_relatorio_diano(self, dados: Dict) -> str:
        """
        Gera relatorio do dia em Excel.
        
        dados = {
            "data": "2026-06-25",
            "apps_usados": [{"nome": "Chrome", "tempo_min": 120}, ...],
            "comandos_jarvis": 15,
            "memorias_criadas": 3,
            "tarefas_concluidas": 5,
            "tempo_focado_min": 180,
            "resumo_ia": "Dia produtivo..."
        }
        """
        if not OPENPYXL_OK:
            return self._gerar_csv_diano(dados)

        data_str = dados.get("data", datetime.now().strftime("%Y-%m-%d"))
        filename = f"relatorio_dia_{data_str}.xlsx"
        filepath = self.reports_dir / filename

        wb = openpyxl.Workbook()

        # === aba Resumo ===
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        self._estilizar_excel(ws_resumo, f"Relatório Diário - {data_str}")

        headers_resumo = ["Métrica", "Valor"]
        ws_resumo.append(headers_resumo)
        ws_resumo.append(["Data", data_str])
        ws_resumo.append(["Comandos Jarvis", dados.get("comandos_jarvis", 0)])
        ws_resumo.append(["Memórias Criadas", dados.get("memorias_criadas", 0)])
        ws_resumo.append(["Tarefas Concluídas", dados.get("tarefas_concluidas", 0)])
        ws_resumo.append(["Tempo Focado (min)", dados.get("tempo_focado_min", 0)])
        ws_resumo.append(["Tempo Focado (h)", round(dados.get("tempo_focado_min", 0) / 60, 1)])
        ws_resumo.append([])
        ws_resumo.append(["Resumo IA", dados.get("resumo_ia", "Sem resumo")])

        # === aba Apps ===
        ws_apps = wb.create_sheet("Apps Usados")
        self._estilizar_excel(ws_apps, "Apps Mais Usados")

        headers_apps = ["App", "Tempo (min)", "Tempo (h)", "% do Dia"]
        ws_apps.append(headers_apps)
        apps = dados.get("apps_usados", [])
        total_min = sum(a.get("tempo_min", 0) for a in apps)
        for app in apps:
            tempo = app.get("tempo_min", 0)
            horas = round(tempo / 60, 1)
            pct = round((tempo / total_min * 100), 1) if total_min > 0 else 0
            ws_apps.append([app.get("nome", "?"), tempo, horas, f"{pct}%"])

        # Grafico de barras
        if apps and len(apps) > 1:
            chart = BarChart()
            chart.title = "Top Apps por Tempo"
            chart.y_axis.title = "Minutos"
            chart.x_axis.title = "App"
            data_ref = Reference(ws_apps, min_col=2, min_row=1, max_row=len(apps) + 1)
            cats = Reference(ws_apps, min_col=1, min_row=2, max_row=len(apps) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 20
            chart.height = 12
            ws_apps.add_chart(chart, "F2")

        # Salvar
        wb.save(str(filepath))
        logger.info(f"Relatório diário salvo: {filepath}")
        return str(filepath)

    def _gerar_csv_diano(self, dados: Dict) -> str:
        """Fallback CSV se openpyxl nao estiver disponivel."""
        data_str = dados.get("data", datetime.now().strftime("%Y-%m-%d"))
        filename = f"relatorio_dia_{data_str}.csv"
        filepath = self.reports_dir / filename

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Métrica", "Valor"])
            writer.writerow(["Data", data_str])
            writer.writerow(["Comandos Jarvis", dados.get("comandos_jarvis", 0)])
            writer.writerow(["Memórias Criadas", dados.get("memorias_criadas", 0)])
            writer.writerow(["Tarefas Concluídas", dados.get("tarefas_concluidas", 0)])
            writer.writerow(["Tempo Focado (min)", dados.get("tempo_focado_min", 0)])
            writer.writerow([])
            writer.writerow(["App", "Tempo (min)"])
            for app in dados.get("apps_usados", []):
                writer.writerow([app.get("nome", "?"), app.get("tempo_min", 0)])

        logger.info(f"Relatório CSV salvo: {filepath}")
        return str(filepath)

    def exportar_memorias(self, memorias: List[Dict]) -> str:
        """
        Exporta memorias para CSV/Excel.
        
        memorias = [{"timestamp": "...", "categoria": "...", "texto": "..."}, ...]
        """
        if not memorias:
            return "Nenhuma memória para exportar."

        if OPENPYXL_OK:
            return self._exportar_memorias_excel(memorias)
        return self._exportar_memorias_csv(memorias)

    def _exportar_memorias_excel(self, memorias: List[Dict]) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"memorias_export_{timestamp}.xlsx"
        filepath = self.reports_dir / filename

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Memórias"
        self._estilizar_excel(ws, "Exportação de Memórias JARVIS")

        headers = ["Data/Hora", "Categoria", "Texto"]
        ws.append(headers)
        for mem in memorias:
            ws.append([
                mem.get("timestamp", ""),
                mem.get("categoria", ""),
                mem.get("texto", ""),
            ])

        wb.save(str(filepath))
        logger.info(f"Memórias exportadas: {filepath}")
        return str(filepath)

    def _exportar_memorias_csv(self, memorias: List[Dict]) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"memorias_export_{timestamp}.csv"
        filepath = self.reports_dir / filename

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Data/Hora", "Categoria", "Texto"])
            for mem in memorias:
                writer.writerow([
                    mem.get("timestamp", ""),
                    mem.get("categoria", ""),
                    mem.get("texto", ""),
                ])

        logger.info(f"Memórias CSV exportadas: {filepath}")
        return str(filepath)

    def gerar_resumo_semanal(self, dados_semana: Dict) -> str:
        """
        Gera resumo semanal em Excel.
        
        dados_semana = {
            "periodo": "2026-06-19 a 2026-06-25",
            "dias": [{"data": "...", "comandos": 10, "focado_min": 120}, ...],
            "total_comandos": 70,
            "total_focado_min": 840,
            "media_diaria_comandos": 10,
            "media_diaria_focado_h": 2.0,
        }
        """
        if not OPENPYXL_OK:
            return self._gerar_csv_semanal(dados_semana)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"resumo_semanal_{timestamp}.xlsx"
        filepath = self.reports_dir / filename

        wb = openpyxl.Workbook()

        # === aba Resumo ===
        ws = wb.active
        ws.title = "Resumo Semanal"
        periodo = dados_semana.get("periodo", "Semana")
        self._estilizar_excel(ws, f"Resumo Semanal - {periodo}")

        headers = ["Métrica", "Valor"]
        ws.append(headers)
        ws.append(["Período", periodo])
        ws.append(["Total Comandos Jarvis", dados_semana.get("total_comandos", 0)])
        ws.append(["Total Focado (min)", dados_semana.get("total_focado_min", 0)])
        ws.append(["Total Focado (h)", round(dados_semana.get("total_focado_min", 0) / 60, 1)])
        ws.append(["Média Diária Comandos", dados_semana.get("media_diaria_comandos", 0)])
        ws.append(["Média Diária Focado (h)", dados_semana.get("media_diaria_focado_h", 0)])

        # === aba Detalhes por Dia ===
        ws_dias = wb.create_sheet("Detalhes por Dia")
        self._estilizar_excel(ws_dias, "Detalhes Diários")

        headers_dias = ["Data", "Comandos", "Focado (min)", "Focado (h)"]
        ws_dias.append(headers_dias)
        for dia in dados_semana.get("dias", []):
            ws_dias.append([
                dia.get("data", ""),
                dia.get("comandos", 0),
                dia.get("focado_min", 0),
                round(dia.get("focado_min", 0) / 60, 1),
            ])

        # Grafico de barras
        dias = dados_semana.get("dias", [])
        if dias and len(dias) > 1:
            chart = BarChart()
            chart.title = "Foco por Dia"
            chart.y_axis.title = "Minutos"
            data_ref = Reference(ws_dias, min_col=3, min_row=1, max_row=len(dias) + 1)
            cats = Reference(ws_dias, min_col=1, min_row=2, max_row=len(dias) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 20
            chart.height = 12
            ws_dias.add_chart(chart, "F2")

        wb.save(str(filepath))
        logger.info(f"Resumo semanal salvo: {filepath}")
        return str(filepath)

    def _gerar_csv_semanal(self, dados_semana: Dict) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"resumo_semanal_{timestamp}.csv"
        filepath = self.reports_dir / filename

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Data", "Comandos", "Focado (min)"])
            for dia in dados_semana.get("dias", []):
                writer.writerow([
                    dia.get("data", ""),
                    dia.get("comandos", 0),
                    dia.get("focado_min", 0),
                ])

        logger.info(f"Resumo semanal CSV salvo: {filepath}")
        return str(filepath)

    def gerar_planilha_gastos(self, gastos: List[Dict]) -> str:
        """
        Gera planilha de gastos.
        
        gastos = [{"data": "...", "descricao": "...", "valor": 25.50, "categoria": "Alimentação"}, ...]
        """
        if not OPENPYXL_OK:
            return self._gerar_csv_gastos(gastos)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gastos_{timestamp}.xlsx"
        filepath = self.reports_dir / filename

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos"
        self._estilizar_excel(ws, "Controle de Gastos JARVIS")

        headers = ["Data", "Descrição", "Valor (R$)", "Categoria"]
        ws.append(headers)
        total = 0
        for g in gastos:
            valor = g.get("valor", 0)
            total += valor
            ws.append([
                g.get("data", ""),
                g.get("descricao", ""),
                valor,
                g.get("categoria", "Outros"),
            ])

        ws.append([])
        ws.append(["", "TOTAL", total, ""])

        # Estilo para total
        last_row = ws.max_row
        ws.cell(row=last_row, column=2).font = Font(bold=True)
        ws.cell(row=last_row, column=3).font = Font(bold=True, color="FF0000")

        # Grafico por categoria
        if gastos:
            # Agrupar por categoria
            categorias = {}
            for g in gastos:
                cat = g.get("categoria", "Outros")
                categorias[cat] = categorias.get(cat, 0) + g.get("valor", 0)

            ws_cat = wb.create_sheet("Por Categoria")
            ws_cat.append(["Categoria", "Total (R$)"])
            for cat, total_cat in sorted(categorias.items(), key=lambda x: -x[1]):
                ws_cat.append([cat, total_cat])

            if len(categorias) > 1:
                chart = BarChart()
                chart.title = "Gastos por Categoria"
                data_ref = Reference(ws_cat, min_col=2, min_row=1, max_row=len(categorias) + 1)
                cats_ref = Reference(ws_cat, min_col=1, min_row=2, max_row=len(categorias) + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.width = 18
                chart.height = 12
                ws_cat.add_chart(chart, "D2")

        wb.save(str(filepath))
        logger.info(f"Planilha de gastos salva: {filepath}")
        return str(filepath)

    def _gerar_csv_gastos(self, gastos: List[Dict]) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gastos_{timestamp}.csv"
        filepath = self.reports_dir / filename

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Data", "Descrição", "Valor (R$)", "Categoria"])
            total = 0
            for g in gastos:
                valor = g.get("valor", 0)
                total += valor
                writer.writerow([
                    g.get("data", ""),
                    g.get("descricao", ""),
                    valor,
                    g.get("categoria", "Outros"),
                ])
            writer.writerow([])
            writer.writerow(["", "TOTAL", total, ""])

        logger.info(f"CSV de gastos salvo: {filepath}")
        return str(filepath)

    def listar_relatorios(self) -> List[str]:
        """Lista todos os relatorios salvos."""
        relatorios = []
        for f in sorted(self.reports_dir.iterdir(), reverse=True):
            if f.suffix in (".xlsx", ".csv"):
                relatorios.append(f.name)
        return relatorios

    def obter_caminho(self, nome_arquivo: str) -> Optional[str]:
        """Retorna caminho completo de um relatorio."""
        filepath = self.reports_dir / nome_arquivo
        if filepath.exists():
            return str(filepath)
        return None


_generator = None

def get_planilhas_generator():
    global _generator
    if _generator is None:
        _generator = PlanilhasGenerator()
    return _generator
